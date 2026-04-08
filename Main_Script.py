import os
import time
import re
import requests
import pdfplumber
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import queue

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

import tkinter as tk
from tkinter import ttk
from tkcalendar import DateEntry
import datetime

from rapidfuzz import process, fuzz


# -------------------------
# Config
# -------------------------
DOWNLOAD_FOLDER = r"C:\Users\HP\Desktop\NCIT"
MASTER_FILE = r"C:\Users\HP\Desktop\NCIT\Master Database.xlsx"

os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)

MAX_CAPTCHA_ATTEMPTS = 3
MAX_WORKERS = 5


# -------------------------
# LOGGER
# -------------------------
log_queue = queue.Queue()

def log(msg):
    timestamp = datetime.datetime.now().strftime("%H:%M:%S")
    log_queue.put(f"[{timestamp}] {msg}")


# -------------------------
# Helpers
# -------------------------
def extract_date_from_filename(filename):
    match = re.search(r"\d{2}\.\d{2}\.\d{4}", filename)
    return match.group() if match else ""


def extract_against(text):
    if not text:
        return ""

    text = str(text).replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()

    parts = re.split(r'\b(vs\.?|v/s|versus)\b', text, flags=re.IGNORECASE)

    if len(parts) >= 3:
        result = parts[-1].strip()
        result = re.sub(r'^[\.\,\-\:]+', '', result).strip()
        return result

    return ""


def safe_parse_captcha(text):
    tokens = text.strip().replace("=", "").split()
    a = int(re.sub(r"\D", "", tokens[0]))
    b = int(re.sub(r"\D", "", tokens[2]))
    return a + b if "+" in text else a - b


# -------------------------
# Driver
# -------------------------
def start_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.maximize_window()

    return driver


# -------------------------
# PDF Extraction
# -------------------------
def extract_all_rows(pdf_path, pdf_name):

    log(f"Parsing PDF: {pdf_name}")

    rows = []
    date = extract_date_from_filename(pdf_name)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[1:]:
                tables = page.extract_tables()

                for table in tables:
                    for row in table:
                        if not row:
                            continue

                        clean_row = [
                            str(cell).strip() if cell else ""
                            for cell in row
                        ]

                        rows.append([date] + clean_row)

    except Exception as e:
        log(f"Parse error: {e}")

    return rows


def merge_multiline_rows(rows, header_len):
    merged = []
    current = None

    for row in rows:

        if len(row) < header_len:
            row += [""] * (header_len - len(row))
        elif len(row) > header_len:
            row = row[:header_len]

        sr = str(row[1]).strip()
        case_no = str(row[2]).strip()

        is_new = False

        if re.match(r'^\d+', sr):
            is_new = True
        elif case_no:
            is_new = True

        if is_new:
            if current:
                merged.append(current)
            current = row
        else:
            if current:
                for i in range(len(row)):
                    if row[i]:
                        current[i] += "\n" + row[i]

    if current:
        merged.append(current)

    return merged


# -------------------------
# Matching Function
# -------------------------

def run_matching(nclt_file_path):

    log("🔍 Running fuzzy...")

    nclt_data = pd.read_excel(nclt_file_path, sheet_name="NCLT_Data")
    sheet1 = pd.read_excel(MASTER_FILE)

    
    sheet1["EST_NAME_clean"] = sheet1["EST_NAME"].astype(str).str.lower().str.strip()

    est_list = sheet1["EST_NAME_clean"].dropna().tolist()

    def find_match(name):
        match = process.extractOne(name, est_list, scorer=fuzz.token_sort_ratio)
        if match:
            matched_name, score, _ = match
            if score >= 80:
                return pd.Series(["Match", matched_name, score])
        return pd.Series(["No Match", "", 0])

    nclt_data[["Match_Status","Matched_EST_NAME","Similarity_%"]] = \
        nclt_data[["Match_Status","Matched_EST_NAME","Similarity_%"]] = \
    nclt_data["Against"].fillna("").astype(str).str.lower().str.strip().apply(find_match)

    output_path = os.path.join(os.path.dirname(nclt_file_path),"FINAL_OUTPUT.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        nclt_data.to_excel(writer, sheet_name="Matched_Data", index=False)
        sheet1.to_excel(writer, sheet_name="Master_Data", index=False)

    wb = load_workbook(output_path)

    fill = PatternFill(start_color="668cff", end_color="668cff", fill_type="solid")
    font = Font(bold=True)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font

    wb.save(output_path)

    log("✅ Matching complete")
    os.startfile(output_path)


# -------------------------
# Excel Save
# -------------------------
def save_rows_to_excel(all_rows, header, output_path, court_name):

    header = ["Date"] + header
    expected_len = len(header)

    cleaned_rows = merge_multiline_rows(all_rows, expected_len)

    df = pd.DataFrame(cleaned_rows, columns=header)
    df.insert(1, "Court Name", court_name)

    df["Against"] = df["Name of the Parties"].apply(extract_against)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="NCLT_Data", index=False)

    wb = load_workbook(output_path)
    ws = wb["NCLT_Data"]

    fill = PatternFill(start_color="668cff", end_color="668cff", fill_type="solid")
    font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(output_path)

    log(f"Excel saved: {output_path}")
    return output_path


# -------------------------
# PDF Downloader
# -------------------------
def process_pdf(pdf_url, pdf_name, run_folder):

    try:
        pdf_name = "".join(ch for ch in pdf_name if ch not in r'\/:*?"<>|')

        if not pdf_name.lower().endswith(".pdf"):
            pdf_name += ".pdf"

        log(f"Downloading: {pdf_name}")

        save_path = os.path.join(run_folder, pdf_name)

        resp = requests.get(pdf_url, timeout=30)
        resp.raise_for_status()

        with open(save_path, "wb") as f:
            f.write(resp.content)

        return extract_all_rows(save_path, pdf_name)

    except Exception as e:
        log(f"Download error: {e}")
        return []


# -------------------------
# Main Scraper
# -------------------------
def run_nclt(start_date, end_date, run_folder, court_choice):

    driver = start_driver()
    wait = WebDriverWait(driver, 20)

    try:
        log("Opening NCLT site...")

        driver.get("https://nclt.gov.in/all-couse-list")

        Select(wait.until(EC.presence_of_element_located(
            (By.ID, "edit-field-nclt-benches-list-target-id-shs-0-0")
        ))).select_by_visible_text(court_choice)

        start_elem = wait.until(EC.presence_of_element_located((By.ID, "edit-field-cause-date-value")))
        end_elem = wait.until(EC.presence_of_element_located((By.ID, "edit-field-cause-date-value-1")))

        start_elem.send_keys(start_date)
        end_elem.send_keys(end_date)

        ActionChains(driver).move_by_offset(0, 0).click().perform()
        time.sleep(1)

        captcha = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".field-prefix"))).text
        answer = safe_parse_captcha(captcha)

        driver.find_element(By.ID, "edit-captcha-response").send_keys(str(answer))
        driver.find_element(By.ID, "edit-submit-all-couse-list").click()

        log("Captcha solved")

        time.sleep(3)

        pdf_links = driver.find_elements(By.CSS_SELECTOR, "td.views-field-field-upload-file a")
        log(f"Found {len(pdf_links)} PDFs")

        all_rows = []

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = [
                executor.submit(process_pdf, link.get_attribute("href"), link.text.strip(), run_folder)
                for link in pdf_links
            ]

            for future in as_completed(futures):
                all_rows.extend(future.result())

        header = [
            "Sr.","CP/CA/IA/MA No.","CA/IA No.","Purpose","Section/Rule",
            "Name of the Parties","Name of Counsel for Petitioner/ Applicant",
            "Name of Counsel for Respondent",
            "Name of (1) IRP/(2) RP/(3) liquidator","Remarks"
        ]

        if all_rows:
            excel_path = save_rows_to_excel(
                all_rows,
                header,
                os.path.join(run_folder, f"{court_choice}.xlsx"),
                court_choice,
            )

            run_matching(excel_path)

    finally:
        driver.quit()


# -------------------------
# Run Process
# -------------------------
def run_process(start_date, end_date, court_choice):

    run_folder = os.path.join(
        DOWNLOAD_FOLDER,
        datetime.datetime.now().strftime("Run_%Y%m%d_%H%M%S")
    )

    os.makedirs(run_folder, exist_ok=True)

    run_nclt(start_date, end_date, run_folder, court_choice)


# -------------------------
# PROGRESS WINDOW
# -------------------------
def launch_progress_window(start_date, end_date, court_choice):

    win = tk.Tk()
    win.title("Processing...")

    win.geometry("520x320")
    win.attributes("-topmost", True)

    progress = ttk.Progressbar(win, mode="indeterminate", length=450)
    progress.pack(pady=10)
    progress.start(10)

    log_box = tk.Text(win, height=15, bg="black", fg="#00ff90")
    log_box.pack(fill="both", expand=True)

    def update_logs():
        while not log_queue.empty():
            log_box.insert(tk.END, log_queue.get() + "\n")
            log_box.see(tk.END)
        win.after(200, update_logs)

    update_logs()

    def run():
        try:
            run_process(start_date, end_date, court_choice)
            log("✅ Done")
        except Exception as e:
            log(f"Error: {e}")
        finally:
            progress.stop()
            win.after(2000, win.destroy)

    threading.Thread(target=run, daemon=True).start()

    win.mainloop()


# -------------------------
# UI
# -------------------------
def launch_ui():

    root = tk.Tk()
    root.geometry("700x400")
    root.configure(bg="#f0f4f7")
    root.title("NCLT Data Extractor")

    tk.Label(root, text="NCLT Data Extractor", font=("Segoe UI", 16, "bold"),
             bg="#f0f4f7", fg="#2c3e50").pack(pady=15)

    frame = tk.Frame(root, bg="#f0f4f7")
    frame.pack(pady=10)

    tk.Label(frame, text="Start Date:", font=("Segoe UI", 12), bg="#f0f4f7").grid(row=0,column=0,padx=10,pady=10)
    start_cal = DateEntry(frame, date_pattern="mm/dd/yyyy", font=("Segoe UI", 12))
    start_cal.grid(row=0,column=1,padx=10,pady=10)

    tk.Label(frame, text="End Date:", font=("Segoe UI", 12), bg="#f0f4f7").grid(row=1,column=0,padx=10,pady=10)
    end_cal = DateEntry(frame, date_pattern="mm/dd/yyyy", font=("Segoe UI", 12))
    end_cal.grid(row=1,column=1,padx=10,pady=10)

    tk.Label(frame, text="Select Court:", font=("Segoe UI", 12), bg="#f0f4f7").grid(row=2,column=0,padx=10,pady=10)

    court_var = tk.StringVar()

    court_dropdown = ttk.Combobox(
        frame,
        textvariable=court_var,
        values=[
            "Ahmedabad Bench Court-I","Ahmedabad Bench Court-II",
            "Mumbai Bench Court-I","Mumbai Bench Court-II",
            "Mumbai Bench Court-III","Mumbai Bench Court-IV",
            "Mumbai Bench Court-V","Mumbai Bench Court-VI",
        ],
        state="readonly",
        width=28
    )

    court_dropdown.grid(row=2,column=1,padx=10,pady=10)
    court_var.set("Ahmedabad Bench Court-I")

    def start_process_ui():
        start_date_val = start_cal.get()
        end_date_val = end_cal.get()
        court_choice_val = court_var.get()

        root.destroy()

        launch_progress_window(
            start_date_val,
            end_date_val,
            court_choice_val
        )

    tk.Button(root, text="Start Generating", font=("Segoe UI", 12, "bold"),
              bg="#3498db", fg="white", command=start_process_ui).pack(pady=20)

    root.mainloop()


if __name__ == "__main__":
    launch_ui()
